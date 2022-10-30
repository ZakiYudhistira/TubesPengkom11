from cProfile import label
import pathlib
# from profile import _Label
from tkinter import *
from datetime import date, datetime
from PIL import ImageTk, Image
import pyrebase, tkinter
import webbrowser
import openpyxl, xlrd

firebaseConfig={'apiKey': "AIzaSyDR4ZslCMZgrl2O1DDzBgoaspzzTScCYoE",
    'authDomain': "tubes1-6911.firebaseapp.com",
    'databaseURL': "https://tubes1-6911.firebaseio.com",
    'projectId': "tubes1-6911",
    'storageBucket': "tubes1-6911.appspot.com",
    'messagingSenderId': "437693074566",
    'appId': "1:437693074566:web:25f5e30e76bf8a1dccd507",
    'measurementId': "G-Q2YCXF2SYT"
}

firebase = pyrebase.initialize_app(firebaseConfig)
db=firebase.database(); auth=firebase.auth(); storage=firebase.storage()
# file = pathlib.Path('database.xlsx')

def showFrame(frame):
    frame.tkraise()

windowUtama = Tk()
windowUtama.state('zoomed')
# windowUtama.rowconfigure(0, weight=1)
# windowUtama.columnconfigure(0, weight=1)
windowUtama.title("Welcome to Jasamarga Tol")
width = int(windowUtama.winfo_screenwidth()); height = int(windowUtama.winfo_screenheight())
Page1 = Frame(windowUtama, width=width, height=height); Page1.grid(row=0, column=0, sticky='nsew')
Page2 = Frame(windowUtama, width=width, height=height); Page2.grid(row=0, column=0, sticky='nsew')
Page3 = Frame(windowUtama, width=width, height=height); Page3.grid(row=0, column=0, sticky='nsew')
Page4 = Frame(windowUtama, width=width, height=height); Page4.grid(row=0, column=0, sticky='nsew')
Page5 = Frame(windowUtama, width=width, height=height); Page5.grid(row=0, column=0, sticky='nsew')
# icon = Image.open('icon toll.png'); ico = ImageTk.PhotoImage(icon); Page1.wm_iconphoto(False, ico)
windowUtama.configure(bg='#f0f4fa')

showFrame(Page1)

file = openpyxl.load_workbook('database.xlsx')
fileSheet = file.sheetnames
sheet1 = file["user"]
sheet2 = file["log"]

# -- PAGE 1 -- # (Login Signup Page)

def Login():
    success = False
    try:
        global userAcc
        auth.sign_in_with_email_and_password(loginEmail.get(), loginPass.get())
        Label(Page1, text="Successfully logged in!").pack(); userAcc = mailSignin.get()
        sheet2.cell(column=8, row=sheet2.max_row+1, value=userAcc)
        file.save(r'database.xlsx')
        success = True
    except:
        success = False
        failLogIn = Label(Page1, text="Either email or password are wrong"); failLogIn.pack()
    if success == True:
        showFrame(Page2)

def Signup():
    try:
        auth.create_user_with_email_and_password(signupEmail.get(), signupPass.get())
        LabelSignedUp = Label(Page1, text="Successfully created an account!"); LabelSignedUp.pack()
        mail = emailSignup.get(); password = passSignup.get(); saldo = saldoSignup.get(); name = nameSignup.get() 
        sheet1.cell(column=2, row=sheet1.max_row+1, value=mail)
        sheet1.cell(column=3, row=sheet1.max_row, value=name)
        sheet1.cell(column=5, row=sheet1.max_row, value=password)
        sheet1.cell(column=6, row=sheet1.max_row, value=saldo)
    except:
        failSignedUp = Label(Page1, text="Sign Up Failed"); failSignedUp.pack()

font = ("Montserrat", 10, "bold")
mailSignin = StringVar(); passSignin = StringVar()
Label(Page1, text="Input ID", font= font).pack()
loginEmail = Entry(Page1, textvariable=mailSignin, width=100); loginEmail.pack();
loginPass = Entry(Page1, width=100); loginPass.pack()
LoginButton = Button(Page1, text="Login", command=Login); LoginButton.pack()

emailSignup = StringVar(); passSignup = StringVar(); saldoSignup = IntVar(); saldoSignup.set(''); nameSignup = StringVar()
Label(Page1, text="Create Account", font=font).pack()
Label(Page1, text="E-Mail", font=("Lato", 12, 'bold')).pack()
signupEmail = Entry(Page1, textvariable=emailSignup, width=100); signupEmail.pack()
Label(Page1, text="Nama", font=("Lato", 12, 'bold')).pack()
signupName = Entry(Page1, textvariable=nameSignup, width=100); signupName.pack()
Label(Page1, text="Password", font=("Lato", 12, 'bold')).pack()
signupPass = Entry(Page1, textvariable=passSignup, width=100); signupPass.pack()
Label(Page1, text="Saldo Awal", font=("Lato", 12, 'bold')).pack()
firstSaldo = Entry(Page1, textvariable=saldoSignup, width=100); firstSaldo.pack()
# reviewSignupPass = Entry(Page1, width=100); reviewSignupPass.pack()
SignupButton = Button(Page1, text="Sign Up", command=Signup); SignupButton.pack()

# if reviewSignupPass.get() != signupPass:
#     PassnotSame = Label()

def firebaseconsole():
    webbrowser.open("https://console.firebase.google.com/u/0/project/tubes1-6911/authentication/users")
startButton = Button(Page1, text="Manage", command=firebaseconsole); startButton.pack()
next = Button(Page1, text="Next", command=lambda: showFrame(Page2)); next.pack()

# Label = Label(text= waktuMasuk)
# Label.pack()

#-- ACCOUNT DASHBOARD --#
def openAcc():
    dashboard = Toplevel()
    dashboard.geometry('400x500')
    dashboard.title("Account Dashboard")
    Label(dashboard, text="Account Management", font=("Montserrat", 16, 'bold')).place(x=40, y=10)
    Label(dashboard, text=userAcc).place(x=20, y=30)
    dashboard.mainloop()

# ------ PAGE 2 ----- # (Golongan Kendaraan)


bgcolor = '#f9fbff'
Label(Page2, text="Pilih Golongan Kendaraan", font=("Montserrat", 16, "bold"), fg='#000000', bg=bgcolor, pady=10).place(x=50, y=10)
Button(Page2, text="Next", command=lambda: showFrame(Page3)).place(x=50, y=100)
accPhoto = PhotoImage(file='image/Akun.png')
user = Button(Page2, image=accPhoto, command=openAcc); user.place(x=90, y=10)


# ------ PAGE 3 ----- # (Gerbang Masuk)

waktu1 = datetime.now()
waktuMasuk = waktu1.strftime("%H:%M:%S")
Label(Page3, text="Gerbang Masuk", font=("Montserrat", 16, "bold"), fg='#000000', bg=bgcolor, pady=10).grid(row=0, column=3)


user = Button(Page3, image=accPhoto, command=openAcc); user.grid(row=0, column=6)
Label(Page3, text="Tol Semarang", font=("Montserrat", 13, 'bold'),  fg='#000000', bg=bgcolor).grid(row = 1, column = 1)

bawenPhoto = PhotoImage(file='image/Bawen.png')
semarangPhoto = PhotoImage(file='image/Semarang.png')
soloPhoto = PhotoImage(file='image/Solo.png')
tbosoPhoto = PhotoImage(file='image/Tambak Oso.png')
tbsumurPhoto = PhotoImage(file='image/Tambak Sumur.png')
juandaPhoto = PhotoImage(file='image/Juanda.png')
serpongPhoto = PhotoImage(file='image/Serpong.png')
tjpriokPhoto = PhotoImage(file='image/Tanjung Priok.png')
tmminiPhoto = PhotoImage(file='image/Taman Mini.png')

bawen = Button(Page3, image=bawenPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); bawen.grid(row=2, column = 1); 
semarang = Button(Page3, image=semarangPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); semarang.grid(row=2, column=2)
solo = Button(Page3, image=soloPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); solo.grid(row=2, column=3)
tboso = Button(Page3, image=tbosoPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tboso.grid(row=4, column=1)
tbsumur = Button(Page3, image=tbsumurPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tbsumur.grid(row=4, column=2)
juanda = Button(Page3, image=juandaPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); juanda.grid(row=4, column=3)
serpong = Button(Page3, image=serpongPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); serpong.grid(row=6, column=1)
tjpriok = Button(Page3, image=tjpriokPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tjpriok.grid(row=6, column=2)
tmmini = Button(Page3, image=tmminiPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tmmini.grid(row=6, column=3)

Label(Page3, text="Tol Semarang", font=("Montserrat", 13, 'bold'),  fg='#000000', bg=bgcolor).grid(row = 1, column = 1)
tol2 = Label(Page3, text="Tol Surabaya", font=("Montserrat", 13, 'bold'),  fg='#000000', bg=bgcolor).grid(row = 3, column = 1)
tol3 = Label(Page3, text="Tol Jakarta", font=("Montserrat", 13, 'bold'),  fg='#000000', bg=bgcolor).grid(row = 5, column = 1)

button = [bawen, semarang, solo, tboso, tbsumur, juanda, serpong, tjpriok, tmmini]; txtButton = ['Bawen', 'Semarang', 'Solo','Tambak Oso','Tambak Sumur','Juanda','Serpong','Tanjung Priok','Taman Mini']
def ClickedIn(indexTolMasuk):
    for i in range(len(button)):
        button[i]['state'] = NORMAL
    button[indexTolMasuk]['state'] = DISABLED
    global entryNumber
    entryNumber = int(indexTolMasuk)

def SubmitIn():
    global gtMasuk
    for i in range(len(button)):
        if i == entryNumber:
            # hasil = Label(Page3, text=txtButton[i]); hasil.grid(row=9, column=5)
            # hasil = Label(Page2, textvariable=v, text=txtButton[i]); hasil.grid(row=7, column=5)
            gtMasuk = str(txtButton[i]); 
            sheet2.cell(column=1, row=sheet2.max_row, value=waktuMasuk)
            sheet2.cell(column=2, row=sheet2.max_row, value=gtMasuk)
            file.save(r'database.xlsx')
    if entryNumber != '':
        showFrame(Page4)

bawen.config(command=lambda:ClickedIn(0))
semarang.config(command=lambda:ClickedIn(1))
solo.config(command=lambda:ClickedIn(2))
tboso.config(command=lambda:ClickedIn(3))
tbsumur.config(command=lambda:ClickedIn(4))
juanda.config(command=lambda:ClickedIn(5))
serpong.config(command=lambda:ClickedIn(6))
tjpriok.config(command=lambda:ClickedIn(7))
tmmini.config(command=lambda:ClickedIn(8))

submit = Button(Page3, text="Submit", command=SubmitIn); submit.grid(row=8, column=5)


# -- PAGE 4 -- # (Gerbang Keluar)

waktu2 = datetime.now()
waktuKeluar = waktu2.strftime("%H:%M:%S")


Label(Page4, text="Gerbang Keluar", font=("Montserrat", 16, "bold"), fg='#000000', bg=bgcolor, pady=10).place(x=50, y=10)
# Label(Page4, text=gtMasuk).place(x=10, y=100)
Button(Page4, text="Next").place(x=50, y=100)

bawen2 = Button(Page4, image=bawenPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); bawen2.grid(row=2, column = 1); 
semarang2 = Button(Page4, image=semarangPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); semarang2.grid(row=2, column=2)
solo2 = Button(Page4, image=soloPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); solo2.grid(row=2, column=3)
tboso2 = Button(Page4, image=tbosoPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tboso2.grid(row=4, column=1)
tbsumur2 = Button(Page4, image=tbsumurPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tbsumur2.grid(row=4, column=2)
juanda2 = Button(Page4, image=juandaPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); juanda2.grid(row=4, column=3)
serpong2 = Button(Page4, image=serpongPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); serpong2.grid(row=6, column=1)
tjpriok2 = Button(Page4, image=tjpriokPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tjpriok2.grid(row=6, column=2)
tmmini2 = Button(Page4, image=tmminiPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tmmini2.grid(row=6, column=3)

button2 = [bawen2, semarang2, solo2, tboso2, tbsumur2, juanda2, serpong2, tjpriok2, tmmini2]; txtButton2 = ['Bawen', 'Semarang', 'Solo','Tambak Oso','Tambak Sumur','Juanda','Serpong','Tanjung Priok','Taman Mini']

def ClickedOut(indexTolKeluar):
    for i in range(len(button2)):
        button2[i]['state'] = NORMAL
    button2[indexTolKeluar]['state'] = DISABLED
    global exitNumber
    exitNumber = int(indexTolKeluar)

def SubmitExit():
    for i in range(len(button2)):
        if i == exitNumber:
            # hasil = Label(Page3, text=txtButton[i]); hasil.grid(row=9, column=5)
            # hasil = Label(Page2, textvariable=v, text=txtButton[i]); hasil.grid(row=7, column=5)
            global gtKeluar
            gtKeluar = str(txtButton2[i])
            sheet2.cell(column=4, row=sheet2.max_row, value=waktuKeluar)
            sheet2.cell(column=5, row=sheet2.max_row, value=gtKeluar)
            # sheet2.cell(column=9, row=sheet2.max_row, value=userAcc)
            
            file.save(r'database.xlsx')
    if exitNumber != '':
        showFrame(Page5)

bawen2.config(command=lambda:ClickedOut(0))
semarang2.config(command=lambda:ClickedOut(1))
solo2.config(command=lambda:ClickedOut(2))
tboso2.config(command=lambda:ClickedOut(3))
tbsumur2.config(command=lambda:ClickedOut(4))
juanda2.config(command=lambda:ClickedOut(5))
serpong2.config(command=lambda:ClickedOut(6))
tjpriok2.config(command=lambda:ClickedOut(7))
tmmini2.config(command=lambda:ClickedOut(8))

submit2 = Button(Page4, text="Submit", command=SubmitExit); submit2.grid(row=8, column=5)


# -- PAGE 5 -- # (Saldo dan Hasil Perjalanan)

gerbangMasuk = sheet2['B' + str(sheet2.max_row)].value
gerbangKeluar = sheet2['E' + str(sheet2.max_row)].value


Semarang_tol = (["Bawen",23.1],
                ["Semarang",0],
                ["Solo",40])
Jakarta_tol = ( ["Tanjung Priok",12.1],
                ["Serpong",10.1],
                ["Taman Mini",4.5])
Surabaya_tol =( ["Juanda",12.8],
                ["Tambak Sumur",5],
                ["Tambak Oso",9])
Semarang = False
Jakarta = False
Surabaya = False
jarak = 0
for i in range(3):
    if gerbangMasuk == Semarang_tol[i][0]:
        jarak = jarak + Semarang_tol[i][1]
        sheet2.cell(column=3, row=sheet2.max_row, value=Semarang_tol[i][1])
        Semarang = True
    if gerbangMasuk == Jakarta_tol[i][0]:
        jarak = jarak + Jakarta_tol[i][1]
        sheet2.cell(column=3, row=sheet2.max_row, value=Jakarta_tol[i][1])
        Jakarta = True
    if gerbangMasuk == Surabaya_tol[i][0]:
        jarak = jarak + Surabaya_tol[i][1]
        sheet2.cell(column=3, row=sheet2.max_row, value=Surabaya_tol[i][1])
        Surabaya = True
for j in range(3):
    if gerbangKeluar == Semarang_tol[j][0]:
        jarak = jarak + Semarang_tol[j][1]
        sheet2.cell(column=6, row=sheet2.max_row, value=Semarang_tol[i][1])
        Semarang = True
    if gerbangKeluar == Jakarta_tol[j][0]:
        jarak = jarak + Jakarta_tol[j][1]
        sheet2.cell(column=6, row=sheet2.max_row, value=Jakarta_tol[i][1])
        Jakarta = True
    if gerbangKeluar == Surabaya_tol[j][0]:
        jarak = jarak + Surabaya_tol[j][1]
        sheet2.cell(column=6, row=sheet2.max_row, value=Surabaya_tol[i][1])
        Surabaya = True
if Surabaya == True and Jakarta == True:
    jarak = jarak + 797
elif Surabaya == True and Semarang == True:
    jarak = jarak + 351
elif Jakarta == True and Semarang == True:
    jarak = jarak + 446
# print(f"Total jarak yang ditempuh adalah {jarak}")
sheet2.cell(column=8, row=sheet2.max_row, value=jarak)



file.save(r'database.xlsx')


windowUtama.mainloop()