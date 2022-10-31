from cProfile import label
import pathlib
# from profile import _Label
from tkinter import *
from datetime import date, datetime
from PIL import ImageTk, Image
import pyrebase, tkinter
import openpyxl
import xlrd

firebaseConfig={'apiKey': "AIzaSyDR4ZslCMZgrl2O1DDzBgoaspzzTScCYoE",
    'authDomain': "tubes1-6911.firebaseapp.com",
    'databaseURL': "https://tubes1-6911.firebaseio.com",
    'projectId': "tubes1-6911",
    'storageBucket': "tubes1-6911.appspot.com",
    'messagingSenderId': "437693074566",
    'appId': "1:437693074566:web:25f5e30e76bf8a1dccd507",
    'measurementId': "G-Q2YCXF2SYT"
}

firebase = pyrebase.initialize_app(firebaseConfig)
db=firebase.database(); auth=firebase.auth(); storage=firebase.storage()
# file = pathlib.Path('database.xlsx')

def showFrame(frame):
    frame.tkraise()

windowUtama = Tk()
windowUtama.state('zoomed')
windowUtama.title("Welcome to Jasamarga Tol")
width = int(windowUtama.winfo_screenwidth()); height = int(windowUtama.winfo_screenheight())
Page1 = Frame(windowUtama, width=width, height=height); Page1.grid(row=0, column=0, sticky='nsew')
Page2 = Frame(windowUtama, width=width, height=height); Page2.grid(row=0, column=0, sticky='nsew')
Page3 = Frame(windowUtama, width=width, height=height); Page3.grid(row=0, column=0, sticky='nsew')
Page4 = Frame(windowUtama, width=width, height=height); Page4.grid(row=0, column=0, sticky='nsew')
Page5 = Frame(windowUtama, width=width, height=height); Page5.grid(row=0, column=0, sticky='nsew')
# icon = Image.open('icon toll.png'); ico = ImageTk.PhotoImage(icon); Page1.wm_iconphoto(False, ico)
windowUtama.configure(bg='#f0f4fa')

showFrame(Page1)

file = openpyxl.load_workbook('database.xlsx')
fileSheet = file.sheetnames
sheet1 = file["database user"]
sheet2 = file["log"]

# -- PAGE 1 -- # (Login Signup Page)

def Login():
    success = False
    try:
        global userAcc
        auth.sign_in_with_email_and_password(loginEmail.get(), loginPass.get())
        Label(Page1, text="Successfully logged in!").pack(); userAcc = mailSignin.get()
        sheet2.cell(column=8, row=sheet2.max_row+1, value=userAcc)
        file.save(r'database.xlsx')
        success = True
    except:
        success = False
        failLogIn = Label(Page1, text="Either email or password are wrong"); failLogIn.pack()
    if success == True:
        showFrame(Page2)

def Signup():
    try:
        auth.create_user_with_email_and_password(signupEmail.get(), signupPass.get())
        LabelSignedUp = Label(Page1, text="Successfully created an account!"); LabelSignedUp.pack()
        mail = emailSignup.get(); password = passSignup.get(); saldoDaftar = saldoSignup.get(); name = nameSignup.get() 
        sheet1.cell(column=1, row=sheet1.max_row+1, value=mail)
        sheet1.cell(column=2, row=sheet1.max_row, value=name)
        sheet1.cell(column=4, row=sheet1.max_row, value=password)
        sheet1.cell(column=5, row=sheet1.max_row, value=saldoDaftar)
        file.save(r'database.xlsx')
        emailSignup.set(''); passSignup.set(''); saldoDaftar = saldoSignup.set(''); name = nameSignup.set('') 
    except:
        failSignedUp = Label(Page1, text="Sign Up Failed", width=100); failSignedUp.pack()

font = ("Montserrat", 14, "bold")
mailSignin = StringVar(); passSignin = StringVar()
Label(Page1, text="Sign In", font= font, pady=(5)).pack()
Label(Page1, text="E-mail", font=("Calibri", 11), pady=10).pack()
loginEmail = Entry(Page1, textvariable=mailSignin, width=100); loginEmail.pack()
Label(Page1, text="Password", font=("Calibri", 11), pady=10).pack()
loginPass = Entry(Page1, width=100); loginPass.pack()
LoginButton = Button(Page1, text="Login", command=Login); LoginButton.pack()

emailSignup = StringVar(); passSignup = StringVar(); saldoSignup = IntVar(); saldoSignup.set(''); nameSignup = StringVar()
Label(Page1, text="Create Account", font=font, pady=(10)).pack()
Label(Page1, text="E-Mail", font=("Calibri", 11), pady=10).pack()
signupEmail = Entry(Page1, textvariable=emailSignup, width=100); signupEmail.pack()
Label(Page1, text="Nama", font=("Calibri", 11), pady=10).pack()
signupName = Entry(Page1, textvariable=nameSignup, width=100); signupName.pack()
Label(Page1, text="Password", font=("Calibri", 11), pady=10).pack()
signupPass = Entry(Page1, textvariable=passSignup, width=100); signupPass.pack()
Label(Page1, text="Saldo Awal", font=("Calibri", 11), pady=10).pack()
firstSaldo = Entry(Page1, textvariable=saldoSignup, width=100); firstSaldo.pack()
# reviewSignupPass = Entry(Page1, width=100); reviewSignupPass.pack()
SignupButton = Button(Page1, text="Sign Up", command=Signup); SignupButton.pack()

# if reviewSignupPass.get() != signupPass:
#     PassnotSame = Label()

#-- ACCOUNT DASHBOARD --#


def openAcc():
    dashboard = Toplevel(windowUtama)
    dashboard.geometry('400x500')
    dashboard.title("Account Dashboard")
    Label(dashboard, text="Account Management", font=("Montserrat", 16, 'bold')).place(x=40, y=10)
    Label(dashboard, text="Email", font=("Calibri", 12, 'bold')).place(x=40, y=70)
    Label(dashboard, text="Name", font=("Calibri", 12, 'bold')).place(x=40, y=110)
    Label(dashboard, text="Saldo", font=("Calibri", 12, 'bold')).place(x=40, y=150)
    Label(dashboard, text="Histori", font=("Calibri", 12, 'bold')).place(x=40, y=190)
    for row in range(1, sheet1.max_row+1):
        if userAcc == sheet1['A' + str(row)].value:
            Label(dashboard, text=sheet1['B' + str(row)].value, font=("Calibri", 12), padx=10, bg='#f9fbff').place(x=100, y=110)
            saldo = Label(dashboard, text=sheet1['E' + str(row)].value, font=("Calibri", 12), padx=10, bg='#f9fbff'); saldo.place(x=100, y=150)
    count=0
    Label(dashboard, text=userAcc, font=("Calibri", 12), padx = 10, bg='#f9fbff').place(x=100, y=70)
    for row in range(1, sheet2.max_row):
        if userAcc == sheet2['H' + str(row)].value:
            Label(dashboard, text=sheet2['B' + str(row)].value, padx = 5, bg='#f9fbff').place(x=100, y=190+(count*30))
            count+=1
    def windowRefill():
        def isiSaldo():
            for row in range(1, sheet1.max_row+1):
                if userAcc == sheet1['A' + str(row)].value:
                    sheet1['E' + str(row)].value += int(entrySaldo.get())
                    file.save(r'database.xlsx')
                    saldo = Label(dashboard, text=sheet1['E' + str(row)].value, font=("Calibri", 12), padx=10, bg='#f9fbff'); saldo.place(x=100, y=150)
                    refill.destroy()
        refill = Toplevel(dashboard)
        refill.geometry('100x100')
        Label(refill, text="ISI ULANG").pack(side=TOP)
        entrySaldo = Entry(refill, width=75); entrySaldo.pack()
        Button(refill, text="ISI", command=isiSaldo).pack(side=BOTTOM)
        refill.mainloop()
    Button(dashboard, text="ISI SALDO", command=windowRefill).place(x=250, y=130)
    

    dashboard.mainloop()


# ------ PAGE 2 ----- # (Golongan Kendaraan)


# Tarif Jakarta - Semarang : I=800 II=1200 III=1600
# Tarif Semarang - Surabaya : I = 950 II=1450 III=1900
#Tarif subtol Jakarta = 300
#Tarif subtol Semarang = 400
#Tarif subtol Surabaya =450

bgcolor = '#f9fbff'
Label(Page2, text="Pilih Golongan Kendaraan", font=("Montserrat", 16, "bold"), fg='#000000', pady=10).place(x=50, y=20)
accPhoto = PhotoImage(file='image/Akun.png')
user = Button(Page2, image=accPhoto, command=openAcc); user.place(x=width-100, y=50)
fotogol1 = PhotoImage(file="image/golongan1.png")
fotogol2 = PhotoImage(file="image/golongan2.png")
fotogol3 = PhotoImage(file="image/golongan3.png")
gol1 = Button(Page2, image=fotogol1, state=NORMAL)
gol2 = Button(Page2, image=fotogol2, state=NORMAL)
gol3 = Button(Page2, image=fotogol3, state=NORMAL)
arrGolongan = [gol1, gol2, gol3]
harGolonganJS = [800, 1200, 1600]
harGolonganSS = [950, 1450, 1900]

def ClickedGol(indexGolongan):
    for i in range(len(arrGolongan)):
        arrGolongan[i]['state'] = NORMAL
    arrGolongan[indexGolongan]['state'] = DISABLED
    global golNumber
    golNumber = int(indexGolongan)

def SubmitGol():
    global golKendaraan
    sheet2.cell(column=9, row=sheet2.max_row, value=golNumber+1)
    file.save(r'database.xlsx')
    if golNumber != '':
        showFrame(Page3)

gol1.config(command=lambda:ClickedGol(0))
gol2.config(command=lambda:ClickedGol(1))
gol3.config(command=lambda:ClickedGol(2))
for i in range(len(arrGolongan)):
    arrGolongan[i].place(x=100+(i*150), y=100)

Button(Page2, text="Next", command=SubmitGol).place(x=50, y=250)

# ------ PAGE 3 ----- # (Gerbang Masuk)

waktu1 = datetime.now()
waktuMasuk = waktu1.strftime("%H:%M:%S")
Label(Page3, text="Gerbang Masuk", font=("Montserrat", 16, "bold"), fg='#000000', bg=bgcolor, pady=10).place(x=width/2-100, y=40)


user = Button(Page3, image=accPhoto, command=openAcc); user.grid(row=0, column=6)

bawenPhoto = PhotoImage(file='image/Bawen.png')
semarangPhoto = PhotoImage(file='image/Semarang.png')
soloPhoto = PhotoImage(file='image/Solo.png')
tbosoPhoto = PhotoImage(file='image/Tambak Oso.png')
tbsumurPhoto = PhotoImage(file='image/Tambak Sumur.png')
juandaPhoto = PhotoImage(file='image/Juanda.png')
serpongPhoto = PhotoImage(file='image/Serpong.png')
tjpriokPhoto = PhotoImage(file='image/Tanjung Priok.png')
tmminiPhoto = PhotoImage(file='image/Taman Mini.png')

bawen = Button(Page3, image=bawenPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); bawen.grid(row=2, column = 1); 
semarang = Button(Page3, image=semarangPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); semarang.grid(row=2, column=2)
solo = Button(Page3, image=soloPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); solo.grid(row=2, column=3)
tboso = Button(Page3, image=tbosoPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tboso.grid(row=4, column=1)
tbsumur = Button(Page3, image=tbsumurPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tbsumur.grid(row=4, column=2)
juanda = Button(Page3, image=juandaPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); juanda.grid(row=4, column=3)
serpong = Button(Page3, image=serpongPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); serpong.grid(row=6, column=1)
tjpriok = Button(Page3, image=tjpriokPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tjpriok.grid(row=6, column=2)
tmmini = Button(Page3, image=tmminiPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tmmini.grid(row=6, column=3)

Label(Page3, text="Tol Semarang", font=("Montserrat", 13, 'bold'),  fg='#000000', bg=bgcolor).grid(row = 1, column = 1)
tol2 = Label(Page3, text="Tol Surabaya", font=("Montserrat", 13, 'bold'),  fg='#000000', bg=bgcolor).grid(row = 3, column = 1)
tol3 = Label(Page3, text="Tol Jakarta", font=("Montserrat", 13, 'bold'),  fg='#000000', bg=bgcolor).grid(row = 5, column = 1)


Semarang_tol = (["Bawen",23.1],
                ["Semarang",0],
                ["Solo",40])
Surabaya_tol =( ["Tambak Oso",9],
                ["Tambak Sumur",5],
                ["Juanda",12.8])
Jakarta_tol = ( ["Serpong",10.1],
                ["Tanjung Priok",12.1],
                ["Taman Mini",4.5])

button = [bawen, semarang, solo, tboso, tbsumur, juanda, serpong, tjpriok, tmmini]; txtButton = ['Bawen', 'Semarang', 'Solo','Tambak Oso','Tambak Sumur','Juanda','Serpong','Tanjung Priok','Taman Mini']
def ClickedIn(indexTolMasuk):
    for i in range(len(button)):
        button[i]['state'] = NORMAL
    button[indexTolMasuk]['state'] = DISABLED
    global entryNumber
    entryNumber = int(indexTolMasuk)

def SubmitIn():
    global gtMasuk
    for i in range(len(button)):
        if i == entryNumber:
            if i < 3:
                kmMasuk = Semarang_tol[i%3][1]
                sheet2.cell(column=10, row=sheet2.max_row, value="Semarang")
                tarif = Semarang_tol[i%3][1]*400
            elif 3 <= i < 6:
                kmMasuk = Surabaya_tol[i%3][1]
                sheet2.cell(column=10, row=sheet2.max_row, value="Surabaya")
                tarif = Surabaya_tol[i%3][1]*450
            elif 6 <= i < 9:
                kmMasuk = Jakarta_tol[i%3][1]
                sheet2.cell(column=10, row=sheet2.max_row, value="Jakarta")
                tarif = Jakarta_tol[i%3][1]*300
            gtMasuk = str(txtButton[i])
            sheet2.cell(column=12, row=sheet2.max_row, value=tarif)
            sheet2.cell(column=1, row=sheet2.max_row, value=waktuMasuk)
            sheet2.cell(column=2, row=sheet2.max_row, value=gtMasuk)
            sheet2.cell(column=3, row=sheet2.max_row, value=kmMasuk)
            file.save(r'database.xlsx')
    if entryNumber != '':
        showFrame(Page4)

bawen.config(command=lambda:ClickedIn(0))
semarang.config(command=lambda:ClickedIn(1))
solo.config(command=lambda:ClickedIn(2))
tboso.config(command=lambda:ClickedIn(3))
tbsumur.config(command=lambda:ClickedIn(4))
juanda.config(command=lambda:ClickedIn(5))
serpong.config(command=lambda:ClickedIn(6))
tjpriok.config(command=lambda:ClickedIn(7))
tmmini.config(command=lambda:ClickedIn(8))

submit = Button(Page3, text="Submit", command=SubmitIn); submit.grid(row=8, column=5)


# -- PAGE 4 -- # (Gerbang Keluar)

waktu2 = datetime.now()
waktuKeluar = waktu2.strftime("%H:%M:%S")


Label(Page4, text="Gerbang Keluar", font=("Montserrat", 16, "bold"), fg='#000000', bg=bgcolor, pady=10).place(x=50, y=10)
# Label(Page4, text=gtMasuk).place(x=10, y=100)

Label(Page4, text="Tol Semarang", font=("Montserrat", 13, 'bold'),  fg='#000000', bg=bgcolor).grid(row = 1, column = 1)
tol22 = Label(Page4, text="Tol Surabaya", font=("Montserrat", 13, 'bold'),  fg='#000000', bg=bgcolor).grid(row = 3, column = 1)
tol32 = Label(Page4, text="Tol Jakarta", font=("Montserrat", 13, 'bold'),  fg='#000000', bg=bgcolor).grid(row = 5, column = 1)

bawen2 = Button(Page4, image=bawenPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); bawen2.grid(row=2, column = 1); 
semarang2 = Button(Page4, image=semarangPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); semarang2.grid(row=2, column=2)
solo2 = Button(Page4, image=soloPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); solo2.grid(row=2, column=3)
tboso2 = Button(Page4, image=tbosoPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tboso2.grid(row=4, column=1)
tbsumur2 = Button(Page4, image=tbsumurPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tbsumur2.grid(row=4, column=2)
juanda2 = Button(Page4, image=juandaPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); juanda2.grid(row=4, column=3)
serpong2 = Button(Page4, image=serpongPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); serpong2.grid(row=6, column=1)
tjpriok2 = Button(Page4, image=tjpriokPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tjpriok2.grid(row=6, column=2)
tmmini2 = Button(Page4, image=tmminiPhoto, borderwidth=0, bg=bgcolor, padx=5, state=NORMAL); tmmini2.grid(row=6, column=3)

button2 = [bawen2, semarang2, solo2, tboso2, tbsumur2, juanda2, serpong2, tjpriok2, tmmini2]; txtButton2 = ['Bawen', 'Semarang', 'Solo','Tambak Oso','Tambak Sumur','Juanda','Serpong','Tanjung Priok','Taman Mini']

def ClickedOut(indexTolKeluar):
    for i in range(len(button2)):
        button2[i]['state'] = NORMAL
    button2[indexTolKeluar]['state'] = DISABLED
    global exitNumber
    exitNumber = int(indexTolKeluar)

def SubmitExit():
    global gtKeluar
    for i in range(len(button2)):
        if i == exitNumber:
            if i < 3:
                kmKeluar = Semarang_tol[i%3][1]
                sheet2.cell(column=11, row=sheet2.max_row, value="Semarang")
                tarif = sheet2.cell(column=12, row=sheet2.max_row).value + Semarang_tol[i%3][1]*400
            elif 3 <= i < 6:
                kmKeluar = Surabaya_tol[i%3][1]
                sheet2.cell(column=11, row=sheet2.max_row, value="Surabaya")
                tarif = sheet2.cell(column=12, row=sheet2.max_row).value + Surabaya_tol[i%3][1]*450
            elif 6 <= i < 9:
                kmKeluar = Jakarta_tol[i%3][1]
                sheet2.cell(column=11, row=sheet2.max_row, value="Jakarta")
                tarif = sheet2.cell(column=12, row=sheet2.max_row).value + Jakarta_tol[i%3][1]*300
            gtKeluar = str(txtButton2[i])
            sheet2.cell(column=12, row=sheet2.max_row, value=tarif)
            sheet2.cell(column=4, row=sheet2.max_row, value=waktuKeluar)
            sheet2.cell(column=5, row=sheet2.max_row, value=gtKeluar)
            sheet2.cell(column=6, row=sheet2.max_row, value=kmKeluar)
            file.save(r'database.xlsx')
    if exitNumber != '':
        In = sheet2.cell(column=10, row=sheet2.max_row).value; Out = sheet2.cell(column=11, row=sheet2.max_row).value
        jarak = sheet2.cell(column=6, row=sheet2.max_row).value + sheet2.cell(column=3, row=sheet2.max_row).value
        tarifAkhir = sheet2.cell(column=12, row=sheet2.max_row).value #sudah biaya entry dan exit
        golAkhir = sheet2.cell(column=9, row=sheet2.max_row).value
        if In == 'Jakarta' and Out == 'Surabaya' or In == 'Surabaya' and Out == 'Jakarta':
            jarak += 797
            if golAkhir == 1:
                tarifAkhir += (351*950) + (446*800)
            elif golAkhir == 2:
                tarifAkhir += (351*1450) + (446*1200)
            elif golAkhir == 3:
                tarifAkhir += (351*1900) + (446*1200)
        if In == 'Semarang' and Out == 'Surabaya' or In == 'Surabaya' and Out == 'Semarang':
            jarak += 351
            if golAkhir == 1:
                tarifAkhir += (351*950)
            elif golAkhir == 2:
                tarifAkhir += (351*1450)
            elif golAkhir == 3:
                tarifAkhir += (351*1900)
        if In == 'Jakarta' and Out == 'Semarang' or In == 'Semarang' and Out == 'Jakarta':
            jarak += 446
            if golAkhir == 1:
                tarifAkhir += (446*800)
            elif golAkhir == 2:
                tarifAkhir += (446*1200)
            elif golAkhir == 3:
                tarifAkhir += (446*1900)
        sheet2.cell(column=7, row=sheet2.max_row, value=jarak)
        sheet2.cell(column=12, row=sheet2.max_row, value=tarifAkhir)
        file.save(r'database.xlsx')
        Label(Page5, text="Rp." + str(tarifAkhir), font=("Montserrat", 50, 'bold')).place(x=180, y=170)
        Label(Page5, text=jarak, font=("Lato", 15, 'bold')).place(x=400, y=270)
        showFrame(Page5)


bawen2.config(command=lambda:ClickedOut(0))
semarang2.config(command=lambda:ClickedOut(1))
solo2.config(command=lambda:ClickedOut(2))
tboso2.config(command=lambda:ClickedOut(3))
tbsumur2.config(command=lambda:ClickedOut(4))
juanda2.config(command=lambda:ClickedOut(5))
serpong2.config(command=lambda:ClickedOut(6))
tjpriok2.config(command=lambda:ClickedOut(7))
tmmini2.config(command=lambda:ClickedOut(8))



submit2 = Button(Page4, text="Submit", command=SubmitExit); submit2.grid(row=8, column=5)

# -- PAGE 5 -- # (Saldo dan Hasil Perjalanan)
Label(Page5, text="Biaya Perjalanan", font=("Montserrat", 60, 'bold')).place(x=180, y=70)
Label(Page5, text="Jarak: ", font=("Montserrat", 15, 'bold')).place(x=180, y=270)
user = Button(Page5, image=accPhoto, command=openAcc); user.place(x=width-70, y=20)

def Bayar():
    for row in range(1, sheet1.max_row+1):
        if userAcc == sheet1['A' + str(row)].value:
            
            saldo = sheet1['E' + str(row)].value 
            pay = sheet2.cell(column=12, row=sheet2.max_row).value
            if saldo < pay:
                Label(Page5, text="Saldo Anda Kurang").place(x=900, y=900)
            else:
                saldo = saldo - pay
                Label(Page5, text="saldo anda sisa " + str(saldo)).place(x=700, y=800)
            sheet1.cell(row=row, column=5, value=saldo)
            file.save(r'database.xlsx')    
    # if pay > dompet:
    #     end.config(state=DISABLED)
    # else:
    #     end.config(state=NORMAL)


end = Button(Page5, text="Submit", font=("Lato", 15, 'bold'), padx=12, pady=3, state=NORMAL, command=Bayar); end.place(x=width/2 - 20, y=height-200)

windowUtama.mainloop()